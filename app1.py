import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from io import BytesIO
import os
from datetime import datetime, date, time, timedelta
import calendar

# Librerías para PDF
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

# ==========================================
# MOTOR DE BASE DE DATOS - GOOGLE SHEETS
# ==========================================
@st.cache_data(ttl=600)
def cargar_competencias_gsheets():
    sheet_id = "1MAIAGFEBerD3Gg-WYfOMvTXQ0uj7JO8UZBag0e8sxjw"
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    try:
        df = pd.read_csv(url)
        df.fillna("", inplace=True) 
        base_datos = {}
        for _, fila in df.iterrows():
            comp = str(fila.iloc[0]).strip()
            rap = str(fila.iloc[1]).strip() if len(fila) > 1 else ""
            if comp != "" and comp.lower() != "nan" and "unnamed" not in comp.lower():
                if comp not in base_datos:
                    base_datos[comp] = []
                if rap != "" and rap.lower() != "nan" and "unnamed" not in rap.lower():
                    base_datos[comp].append(rap)
        base_datos["OTRA (Escribir manualmente)"] = []
        return base_datos
    except Exception as e:
        return {"OTRA (Escribir manualmente)": []}

DB_SENA = cargar_competencias_gsheets()

# ==========================================
# CONFIGURACIÓN DE PÁGINA E INTERFAZ
# ==========================================
st.set_page_config(page_title="Reporte SENA", layout="wide")

col_logo, col_tit = st.columns([1, 5])
with col_logo:
    if os.path.exists("logo_sena.png"):
        st.image("logo_sena.png", width=100)

with col_tit:
    st.markdown("### CENTRO INDUSTRIAL Y DE ENERGIAS ALTERNATIVAS - REGIONAL GUAJIRA")
    st.markdown("#### REPORTE ESTADISTICO DE HORAS MENSUALES INSTRUCTOR")

with st.container():
    c1, c2, c3, c4 = st.columns(4)
    nombre = c1.text_input("Nombre del Instructor")
    cedula = c2.text_input("Cédula")
    meses_str = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    mes = c3.selectbox("Mes del Reporte", meses_str, index=1)
    anio = c4.text_input("Año", value="2026")

# ==========================================
# SECCIÓN 1: HORAS DIRECTAS (FORMACIÓN)
# ==========================================
st.divider()
st.markdown("### 📘 1. Parte Horas Directas - Formación Titulada")

if 'filas' not in st.session_state:
    st.session_state.filas = []

if st.button("➕ Agregar Actividad de Formación"):
    st.session_state.filas.append({
        "ficha": "", "h_inicio": time(8, 0), "h_fin": time(12, 0), 
        "dias": {"L": False, "M": False, "Mi": False, "J": False, "V": False, "S": False},
        "competencia": list(DB_SENA.keys())[0] if DB_SENA else "OTRA (Escribir manualmente)", 
        "rap": "", "horas": 0, 
        "evaluado": "NO", "termino": "NO" 
    })

total_horas_directas = 0

# Diccionario para recordar las fichas que ya se les preguntó el estado
fichas_procesadas = {} 

for i, fila in enumerate(st.session_state.filas):
    with st.expander(f"📌 Ficha: {fila.get('ficha', 'Nueva Ficha')}", expanded=True):
        col_1, col_2, col_3, col_4 = st.columns([1.5, 1, 1, 1])
        fila['ficha'] = col_1.text_input("N° Ficha", value=fila.get('ficha', ''), key=f"f_{i}")
        ficha_actual = fila['ficha'].strip() # Limpiamos espacios
        
        fila['h_inicio'] = col_2.time_input("Hora Inicio", value=fila.get('h_inicio', time(8,0)), key=f"hi_{i}")
        fila['h_fin'] = col_3.time_input("Hora Fin", value=fila.get('h_fin', time(12,0)), key=f"hf_{i}")

        cols_d = st.columns(6)
        nombres_dias = ["L", "M", "Mi", "J", "V", "S"]
        for idx, d_nom in enumerate(nombres_dias):
            fila['dias'][d_nom] = cols_d[idx].checkbox(d_nom, value=fila['dias'].get(d_nom, False), key=f"chk_{d_nom}_{i}")

        mes_num = meses_str.index(mes) + 1
        anio_int = int(anio) if anio.isdigit() else 2026
        _, num_dias_mes = calendar.monthrange(anio_int, mes_num)
        
        mapa_dias = {"L": 0, "M": 1, "Mi": 2, "J": 3, "V": 4, "S": 5}
        dias_seleccionados = [mapa_dias[d] for d in nombres_dias if fila['dias'][d]]
        
        fechas_clase = []
        for dia in range(1, num_dias_mes + 1):
            fecha_actual = date(anio_int, mes_num, dia)
            if fecha_actual.weekday() in dias_seleccionados:
                fechas_clase.append(fecha_actual)

        opciones_fechas = [f.strftime("%d/%m/%Y") for f in fechas_clase]
        fechas_excluidas = st.multiselect("🚫 Fechas a descontar (Festivos, Paro):", opciones_fechas, key=f"excl_{i}")
        
        dt_inicio = datetime.combine(date.today(), fila['h_inicio'])
        dt_fin = datetime.combine(date.today(), fila['h_fin'])
        if dt_fin < dt_inicio: dt_fin += timedelta(days=1)
        horas_diarias = (dt_fin - dt_inicio).seconds / 3600
        
        dias_efectivos = len(fechas_clase) - len(fechas_excluidas)
        horas_totales_actividad = dias_efectivos * horas_diarias
        fila['horas'] = horas_totales_actividad
        total_horas_directas += horas_totales_actividad

        col_4.metric("Total Automático", f"{horas_totales_actividad:g} hrs")

        col_comp, col_rap = st.columns(2)
        lista_comps = list(DB_SENA.keys())
        if not lista_comps: lista_comps = ["OTRA (Escribir manualmente)"]
        fila['competencia'] = col_comp.selectbox("Competencia", lista_comps, key=f"cp_{i}")
        opciones_rap = DB_SENA.get(fila['competencia'], [])
        
        if fila['competencia'] == "OTRA (Escribir manualmente)" or not opciones_rap:
            fila['rap'] = col_rap.text_area("Escriba el RAP", key=f"rpm_{i}")
        else:
            fila['rap'] = col_rap.selectbox("Seleccione el RAP", opciones_rap, key=f"rp_{i}")

        # --- NUEVA LÓGICA INTELIGENTE DE BOTONES ---
        if ficha_actual != "" and ficha_actual in fichas_procesadas:
            # Si la ficha ya la vimos más arriba, copiamos la info y NO mostramos los botones
            fila['evaluado'] = fichas_procesadas[ficha_actual]['evaluado']
            fila['termino'] = fichas_procesadas[ficha_actual]['termino']
            
            st.info(f"💡 Estado heredado del bloque anterior (Ficha {ficha_actual}): **Evaluado:** {fila['evaluado']} | **Terminó:** {fila['termino']}")
        else:
            # Si es la primera vez que vemos la ficha en la lista, preguntamos.
            st.markdown("**Estado de la Competencia / RAP:**")
            col_ev, col_fin = st.columns(2)
            fila['evaluado'] = col_ev.radio("✔️ ¿Está evaluado?", ["SÍ", "NO"], horizontal=True, key=f"ev_{i}", index=1 if fila.get('evaluado', 'NO') == 'NO' else 0)
            fila['termino'] = col_fin.radio("🏁 ¿La competencia terminó?", ["SÍ", "NO"], horizontal=True, key=f"term_{i}", index=1 if fila.get('termino', 'NO') == 'NO' else 0)
            st.markdown("---")
            
            # Guardamos la ficha en la memoria para que los de abajo no pregunten de nuevo
            if ficha_actual != "":
                fichas_procesadas[ficha_actual] = {
                    'evaluado': fila['evaluado'],
                    'termino': fila['termino']
                }


# ==========================================
# SECCIÓN 2: OTRAS ACTIVIDADES (NUEVO)
# ==========================================
st.divider()
st.markdown("### 📙 2. Parte Otras Actividades Instructores Planta")
st.caption("Permisos sindicales, Eventos, Incapacidades, Festivos, etc. (El sistema calcula 8.5 horas por cada día reportado).")

if 'otras_filas' not in st.session_state:
    st.session_state.otras_filas = []

if st.button("➕ Agregar Otra Actividad / Novedad"):
    st.session_state.otras_filas.append({
        "actividad": "Preparación de clases", 
        "f_desde": date.today(), 
        "f_hasta": date.today(),
        "dias": 1.0, 
        "horas": 8.5
    })

total_horas_otras = 0

lista_otras_actividades = [
    "Preparación de clases", "Incapacidad", "Festivo", "Licencia", 
    "Permiso por estudio", "Permiso sindical", "Semana de confraternidad", 
    "Eventos culturales", "Eventos deportivos", "Día del instructor", 
    "Día de la familia", "Reuniones de Centro", "Otra"
]

for j, ofila in enumerate(st.session_state.otras_filas):
    with st.container():
        c_act, c_fd, c_fh, c_dias, c_hr = st.columns([2, 1, 1, 1, 1])
        
        ofila['actividad'] = c_act.selectbox("Actividad", lista_otras_actividades, key=f"oact_{j}")
        ofila['f_desde'] = c_fd.date_input("Fecha Desde", value=ofila['f_desde'], key=f"ofd_{j}")
        ofila['f_hasta'] = c_fh.date_input("Fecha Hasta", value=ofila['f_hasta'], key=f"ofh_{j}")
        ofila['dias'] = c_dias.number_input("Cant. Días (Ej. 4)", min_value=0.0, step=0.5, value=float(ofila.get('dias', 1.0)), key=f"odias_{j}")
        
        ofila['horas'] = ofila['dias'] * 8.5
        c_hr.metric("Total Horas", f"{ofila['horas']:g} hrs")
        
        total_horas_otras += ofila['horas']

# ==========================================
# TOTALES GLOBALES
# ==========================================
total_general_mes = total_horas_directas + total_horas_otras

st.divider()
st.sidebar.markdown("### 🧮 RESUMEN DEL MES")
st.sidebar.metric("Formación Directa", f"{total_horas_directas:g} hrs")
st.sidebar.metric("Otras Actividades", f"{total_horas_otras:g} hrs")
st.sidebar.markdown("---")
st.sidebar.metric("TOTAL GENERAL", f"{total_general_mes:g} hrs")

# ==========================================
# MOTOR DE EXPORTACIÓN A EXCEL
# ==========================================
def crear_excel(datos_formacion, datos_otras, tot_dir, tot_otr, tot_gen):
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Mensual"

    bold = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    fill_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    ws.merge_cells('B2:O2')
    ws['B2'] = "CENTRO INDUSTRIAL Y DE ENERGIAS ALTERNATIVAS - REGIONAL GUAJIRA"
    ws['B2'].font = Font(bold=True, size=12)
    ws['B2'].alignment = center

    ws.merge_cells('B3:O3')
    ws['B3'] = "REPORTE ESTADISTICO DE HORAS MENSUALES INSTRUCTOR"
    ws['B3'].font = Font(bold=True, size=14)
    ws['B3'].alignment = center
    
    ws['B6'] = f"INSTRUCTOR: {nombre}"
    ws['J6'] = f"MES: {mes} / {anio}"
    ws['B7'] = f"CÉDULA: {cedula}"

    ws['B9'] = "PARTE HORAS DIRECTAS - PROGRAMAS EN FORMACIÓN TITULADA"
    ws['B9'].font = bold

    header_labels = ["FICHA", "DESDE", "HASTA", "L", "M", "MI", "J", "V", "S", "COMPETENCIA", "RAP", "EVALUADO", "TERMINÓ", "TOTAL"]
    for col, text in enumerate(header_labels, 2):
        cell = ws.cell(row=10, column=col, value=text)
        cell.font, cell.border, cell.alignment, cell.fill = bold, border, center, fill_header

    current_row = 11
    for f in datos_formacion:
        ws.cell(row=current_row, column=2, value=f['ficha']).border = border
        ws.cell(row=current_row, column=3, value=f['h_inicio'].strftime("%H:%M")).border = border
        ws.cell(row=current_row, column=4, value=f['h_fin'].strftime("%H:%M")).border = border
        
        for d_idx, d_nom in enumerate(["L", "M", "Mi", "J", "V", "S"], 5):
            val = "X" if f['dias'][d_nom] else ""
            c = ws.cell(row=current_row, column=d_idx, value=val)
            c.border, c.alignment = border, center

        ws.cell(row=current_row, column=11, value=f['competencia']).border = border
        ws.cell(row=current_row, column=12, value=f['rap']).border = border
        
        c_ev = ws.cell(row=current_row, column=13, value=f.get('evaluado', 'NO'))
        c_ev.border, c_ev.alignment = border, center
        
        c_fin = ws.cell(row=current_row, column=14, value=f.get('termino', 'NO'))
        c_fin.border, c_fin.alignment = border, center
        
        ws.cell(row=current_row, column=15, value=f['horas']).border = border
        current_row += 1

    ws.cell(row=current_row, column=14, value="TOTAL DIRECTAS:").font = bold
    ws.cell(row=current_row, column=15, value=tot_dir).font = bold
    current_row += 3

    ws.cell(row=current_row, column=2, value="PARTE OTRAS ACTIVIDADES INSTRUCTORES PLANTA").font = bold
    current_row += 1
    
    headers_otras = ["ACTIVIDAD", "F. DESDE", "F. HASTA", "CANT. DÍAS", "HORAS"]
    cols_otras = [2, 10, 11, 13, 14] 
    
    for h_txt, col_num in zip(headers_otras, cols_otras):
        cell = ws.cell(row=current_row, column=col_num, value=h_txt)
        cell.font, cell.border, cell.alignment, cell.fill = bold, border, center, fill_header
    
    current_row += 1
    for of in datos_otras:
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
        c_act = ws.cell(row=current_row, column=2, value=of['actividad'])
        c_act.border, c_act.alignment = border, Alignment(horizontal='left')
        
        ws.cell(row=current_row, column=10, value=of['f_desde'].strftime("%d/%m/%Y")).border = border
        ws.cell(row=current_row, column=11, value=of['f_hasta'].strftime("%d/%m/%Y")).border = border
        ws.cell(row=current_row, column=13, value=of.get('dias', 0)).border = border
        ws.cell(row=current_row, column=14, value=of['horas']).border = border
        current_row += 1
        
    ws.cell(row=current_row, column=13, value="TOTAL OTRAS:").font = bold
    ws.cell(row=current_row, column=14, value=tot_otr).font = bold
    current_row += 2

    ws.cell(row=current_row, column=13, value="TOTAL REPORTADAS:").font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=14, value=tot_gen).font = Font(bold=True, size=12)

    ws.column_dimensions['K'].width = 35
    ws.column_dimensions['L'].width = 45

    wb.save(output)
    return output.getvalue()

# ==========================================
# MOTOR DE EXPORTACIÓN A PDF
# ==========================================
def crear_pdf(datos_formacion, datos_otras, tot_dir, tot_otr, tot_gen):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    
    styles = getSampleStyleSheet()
    style_center = ParagraphStyle(name='Center', parent=styles['Normal'], alignment=1, fontSize=12, spaceAfter=6)
    style_title = ParagraphStyle(name='Title', parent=styles['Heading1'], alignment=1, fontSize=14, spaceAfter=12)
    style_cell = ParagraphStyle(name='Cell', parent=styles['Normal'], fontSize=8, leading=10)
    style_subtitle = ParagraphStyle(name='Sub', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', spaceAfter=6, spaceBefore=10)

    if os.path.exists("logo_sena.png"):
        img = Image("logo_sena.png", width=1*inch, height=1*inch)
        img.hAlign = 'CENTER'
        elements.append(img)
    
    elements.append(Paragraph("<b>CENTRO INDUSTRIAL Y DE ENERGIAS ALTERNATIVAS - REGIONAL GUAJIRA</b>", style_center))
    elements.append(Paragraph("<b>REPORTE ESTADISTICO DE HORAS MENSUALES INSTRUCTOR</b>", style_title))
    
    info_text = f"<b>INSTRUCTOR:</b> {nombre} &nbsp;&nbsp;&nbsp;&nbsp; <b>CÉDULA:</b> {cedula} &nbsp;&nbsp;&nbsp;&nbsp; <b>MES:</b> {mes} / {anio}"
    elements.append(Paragraph(info_text, styles['Normal']))
    
    elements.append(Paragraph("PARTE HORAS DIRECTAS - PROGRAMAS EN FORMACIÓN TITULADA", style_subtitle))
    
    data_table = [["FICHA", "DESDE", "HASTA", "L", "M", "MI", "J", "V", "S", "COMPETENCIA", "RAP", "EVAL", "TERM", "HRS"]]
    for f in datos_formacion:
        row = [
            f['ficha'], f['h_inicio'].strftime("%H:%M"), f['h_fin'].strftime("%H:%M"),
            "X" if f['dias']["L"] else "", "X" if f['dias']["M"] else "",
            "X" if f['dias']["Mi"] else "", "X" if f['dias']["J"] else "",
            "X" if f['dias']["V"] else "", "X" if f['dias']["S"] else "",
            Paragraph(f['competencia'], style_cell),
            Paragraph(f['rap'], style_cell),
            f.get('evaluado', 'NO'), f.get('termino', 'NO'), str(f"{f['horas']:g}")
        ]
        data_table.append(row)
    
    data_table.append(["", "", "", "", "", "", "", "", "", "", "", "", "TOTAL:", str(f"{tot_dir:g}")])

    col_widths_dir = [45, 35, 35, 15, 15, 15, 15, 15, 15, 195, 205, 30, 30, 25]
    t_dir = Table(data_table, colWidths=col_widths_dir)
    t_dir.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (-2,-1), (-1,-1), 'Helvetica-Bold')
    ]))
    elements.append(t_dir)
    
    if datos_otras:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("PARTE OTRAS ACTIVIDADES INSTRUCTORES PLANTA", style_subtitle))
        
        data_otras = [["ACTIVIDAD", "FECHA DESDE", "FECHA HASTA", "CANT. DÍAS", "HRS"]]
        for of in datos_otras:
            data_otras.append([
                of['actividad'], 
                of['f_desde'].strftime("%d/%m/%Y"), 
                of['f_hasta'].strftime("%d/%m/%Y"), 
                str(f"{of.get('dias', 0):g}"),
                str(f"{of['horas']:g}")
            ])
            
        data_otras.append(["", "", "", "TOTAL OTRAS:", str(f"{tot_otr:g}")])
        
        t_otras = Table(data_otras, colWidths=[250, 110, 110, 100, 80])
        t_otras.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('FONTNAME', (-2,-1), (-1,-1), 'Helvetica-Bold')
        ]))
        elements.append(t_otras)
    
    elements.append(Spacer(1, 15))
    elements.append(Paragraph(f"<b>TOTAL HORAS REPORTADAS EN EL MES: &nbsp; <font size=12>{tot_gen:g} Horas</font></b>", style_center))
    
    fecha_descarga = datetime.now().strftime("%d/%m/%Y a las %H:%M")
    style_firma = ParagraphStyle(name='Firma', parent=styles['Normal'], alignment=1, fontSize=11, leading=15)
    
    firma_html = f"""
    <br/><br/><br/><br/>
    _________________________________________________________<br/>
    <b>FIRMA DEL INSTRUCTOR</b><br/>
    {nombre}<br/>
    C.C. {cedula}<br/><br/>
    <i><font size="9" color="gray">Documento generado el: {fecha_descarga}</font></i>
    """
    elements.append(Paragraph(firma_html, style_firma))
    
    doc.build(elements)
    return buffer.getvalue()

# ==========================================
# ÁREA DE DESCARGAS AUTOMÁTICAS
# ==========================================
st.divider()
st.markdown("### 📥 Generación de Reportes Finales")

if not nombre or total_general_mes == 0:
    st.warning("⚠️ Escribe tu nombre y asegúrate de tener horas reportadas para habilitar la descarga.")
else:
    excel_file = crear_excel(st.session_state.filas, st.session_state.otras_filas, total_horas_directas, total_horas_otras, total_general_mes)
    pdf_file = crear_pdf(st.session_state.filas, st.session_state.otras_filas, total_horas_directas, total_horas_otras, total_general_mes)
    
    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        st.download_button(label="🟢 DESCARGAR FORMATO EXCEL COMPLETO", data=excel_file, file_name=f"Reporte_{nombre.replace(' ', '_')}_{mes}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with col_btn2:
        st.download_button(label="🔴 DESCARGAR FORMATO PDF PARA FIRMA", data=pdf_file, file_name=f"Reporte_{nombre.replace(' ', '_')}_{mes}.pdf", mime="application/pdf", use_container_width=True)